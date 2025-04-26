import os
import pandas as pd
from openpyxl import Workbook

def clean_phone_number(phone_str):
    """
    Cleans the phone number:
    1. Converts to string,
    2. Removes spaces and '+' sign,
    3. Removes leading '0' or '90',
    4. Returns the cleaned number if not empty, otherwise an empty string.
    """
    if phone_str is None:
        return ""
    if not isinstance(phone_str, str):
        phone_str = str(phone_str)

    # Remove spaces and '+' sign
    cleaned = phone_str.replace(" ", "").replace("+", "")

    # Remove leading '90'
    if cleaned.startswith("90"):
        cleaned = cleaned[2:]
    # Remove leading '0'
    elif cleaned.startswith("0"):
        cleaned = cleaned[1:]

    # Return empty string if cleaning resulted in an empty string
    if not cleaned:
        print(f"Warning: Phone number '{phone_str}' resulted in an empty string after cleaning. Skipping.")
        return ""

    # Removed the 10-digit check. Now returns the cleaned number regardless of length.
    # Example: If original was ' 123 456', cleaned will be '123456'
    # Example: If original was '+905551112233', cleaned will be '5551112233'
    # Example: If original was '055511122', cleaned will be '55511122'

    return cleaned

def create_directory_if_not_exists(directory_path):
    """
    Create a directory if it does not already exist.

    Args:
        directory_path (str): Path to the directory to create.
    """
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
        print(f"Created directory: {directory_path}")

def read_excel(file_path, phone_col):
    """
    Read an Excel file and return a DataFrame with the specified phone column as string.

    Args:
        file_path (str): Path to the Excel file.
        phone_col (str): Column name for phone numbers.

    Returns:
        pd.DataFrame: DataFrame with the Excel data.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found at '{file_path}'")
        return None

    df = pd.read_excel(file_path, engine='openpyxl', dtype={phone_col: str})
    print(f"Successfully read {len(df)} rows from '{file_path}'.")
    return df

def save_csv(df, output_dir, file_name):
    """
    Save a DataFrame to a CSV file.

    Args:
        df (pd.DataFrame): DataFrame to save.
        output_dir (str): Directory to save the CSV file.
        file_name (str): Name of the CSV file.
    """
    create_directory_if_not_exists(output_dir)
    output_csv_path = os.path.join(output_dir, file_name)
    df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
    print(f"Successfully saved CSV to '{output_csv_path}'.")

def save_excel_with_qr(df, qr_dir, excel_output_path, uuid_column):
    """
    Save a DataFrame to an Excel file with QR codes.

    Args:
        df (pd.DataFrame): DataFrame to save.
        qr_dir (str): Directory containing QR code images.
        excel_output_path (str): Path to save the Excel file.
        uuid_column (str): Column name for UUIDs.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Kodlar"
    ws.column_dimensions['A'].width = 20  # Column width for QR
    # Create header row
    headers = ['qr kodlar', 'ad soyad', 'posta', 'numara']
    ws.append(headers)
    row_number = 2
    for _, row in df.iterrows():
        ws.row_dimensions[row_number].height = 100  # Adjust row height according to QR size
        uuid_value = row.get(uuid_column, '')
        # Construct the image path using UUID, consistent with QRGenerator logic if filenames are based on UUID
        # If QRGenerator uses phone numbers for filenames, this needs adjustment
        # Assuming QRGenerator uses UUID for filenames based on previous context
        # Update: QRGenerator uses phone number, DataExtractor's generate_excel_with_qr uses phone number. Let's correct this function.
        mobile = row.get('mobile', '') # Get the mobile number used for QR filename
        safe_filename = clean_phone_number(mobile) if mobile and str(mobile).strip() else str(uuid_value).strip() # Replicate filename logic
        img_path = os.path.join(qr_dir, f"{safe_filename}.png") # Use the potentially cleaned phone number or UUID as filename base

        # Other columns
        ws.cell(row=row_number, column=2, value=row.get('isim', ''))
        ws.cell(row=row_number, column=3, value=row.get('mail', ''))
        ws.cell(row=row_number, column=4, value=row.get('mobile', ''))
        # Add QR image (if exists)
        if os.path.exists(img_path):
            from openpyxl.drawing.image import Image as OpenpyxlImage
            img = OpenpyxlImage(img_path)
            img.width = 145
            img.height = 145
            ws.add_image(img, f"A{row_number}")
        else:
            ws.cell(row=row_number, column=1, value="No QR")
            print(f"Warning: QR image not found for row {row_number-1} at path: {img_path}") # Added warning
        row_number += 1

    # Adjust other column widths
    for col in ['B', 'C', 'D']:
        max_length = 0
        for cell in ws[col]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col].width = max_length + 5  # add padding

    excel_dir = os.path.dirname(excel_output_path)
    create_directory_if_not_exists(excel_dir)
    wb.save(excel_output_path)
    print(f"Successfully saved Excel to '{excel_output_path}'.")