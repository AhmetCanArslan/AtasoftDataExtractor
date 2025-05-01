import os
import pandas as pd
import uuid
import qrcode
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import subprocess # Added for running the sync script
import sys # Added for getting python executable path
from FileOperations import create_directory_if_not_exists, clean_phone_number # Import clean_phone_number
from QRDesign import overlay_qr_on_template # Removed generate_certificate
from MailSender import send_qr_codes # Removed send_certificates
# Removed FirebaseSync imports related to fetching attendees
from dotenv import load_dotenv
# Removed tqdm import if only used for certificates

# Load environment variables from .env file
load_dotenv()

# --- Configuration ---
# EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH') # Removed: Path will be determined dynamically
INPUT_DIR = 'input' # Define the input directory
PHONE_COLUMN_NAME = os.getenv('PHONE_COLUMN_NAME')
UUID_COLUMN_NAME = os.getenv('UUID_COLUMN_NAME')
COUNTER_COLUMN_NAME = os.getenv('COUNTER_COLUMN_NAME')
CSV_OUTPUT_DIR = os.getenv('CSV_OUTPUT_DIR')
QR_OUTPUT_DIR = os.getenv('QR_OUTPUT_DIR')
# EXCEL_OUTPUT_PATH = os.getenv('EXCEL_OUTPUT_PATH') # Removed: Path will be generated dynamically
FIREBASE_SYNC_SCRIPT = os.getenv('FIREBASE_SYNC_SCRIPT')

# Email configuration
SENDER_EMAIL = os.getenv('SENDER_EMAIL')
SENDER_PASSWORD = os.getenv('SENDER_PASSWORD')
SMTP_SERVER = os.getenv('SMTP_SERVER')
SMTP_PORT = int(os.getenv('SMTP_PORT'))

NAMESPACE = uuid.NAMESPACE_DNS  

# --- Utility Functions ---

def generate_uuid_from_phone(phone_number_str):
    if pd.isna(phone_number_str) or not str(phone_number_str).strip():
        return None
    return str(uuid.uuid5(NAMESPACE, str(phone_number_str)))

# --- Excel Processing & CSV Generation ---
def process_excel(file_path, phone_col, uuid_col, counter_col):
    try:
        if not os.path.exists(file_path):
            print(f"Error: File not found at '{file_path}'")
            return None

        # Load the Excel file, reading the phone column as string
        df = pd.read_excel(file_path, engine='openpyxl', dtype={phone_col: str})
        print(f"Successfully read {len(df)} rows from '{file_path}'.")
        print(f"Columns found: {list(df.columns)}")

        if phone_col not in df.columns:
            print(f"Error: Column '{phone_col}' not found.")
            return None
            
        # Convert phone column to string
        df[phone_col] = df[phone_col].astype(str)

        # Clean phone numbers
        print("Cleaning phone numbers...")
        df['cleaned_phone'] = df[phone_col].apply(clean_phone_number)
        
        # Generate UUID based on cleaned phone numbers
        df[uuid_col] = df['cleaned_phone'].apply(lambda x: generate_uuid_from_phone(x) if x else None)
        print("Generated UUIDs based on cleaned phone numbers.")
        
        # Report the count of rows with invalid/empty phone numbers
        invalid_phones = df[df['cleaned_phone'] == ""].shape[0]
        if invalid_phones > 0:
            print(f"WARNING: {invalid_phones} rows have invalid or empty phone numbers after cleaning.")

        # Add Counter column with initial value zero
        df[counter_col] = 0
        print(f"Added '{counter_col}' column with initial value 0.")

        # Remove the original phone column and use the cleaned number
        df = df.drop(columns=[phone_col])
        df.rename(columns={'cleaned_phone': 'mobile'}, inplace=True)
        print("Replaced original phone column with cleaned phone numbers.")

        # --- Remove unnecessary columns ---
        # Define the exact column names to remove, including multi-line ones
        tc_kimlik_col_name = "TC Kimlik Numarası \n(Bu alanda alınan veriler, ÜNİDES proje kapsamında Gençlik ve Spor Bakanlığı tarafından talep edilmektedir.)"
        kvkk_col_name = "KVKK AYDINLATMA METNİ" # Note: The original Excel might have slightly different spacing/casing
        okul_col_name = "Öğrenim gördüğünüz / mezun olduğunuz öğretim kurumu" # Added for removal
        bolum_col_name = "Öğrenim gördüğünüz / mezun olduğunuz bölüm/dal" # Added for removal
        dogum_tarihi_col_name = "Doğum tarihiniz\n(Bu alanda alınan veriler, ÜNİDES proje kapsamında Gençlik ve Spor Bakanlığı tarafından talep edilmektedir.)" # Added for removal

        columns_to_remove = [
            'Üniversiteniz',
            'Cinsiyet',
            'Bölümünüz',
            'Kaçıncı sınıftasınız? ',
            'Etkinliği nereden duydunuz? ',
            'Etkinliğimizden beklentileriniz nelerdir? ',
            'Eklemek istediğiniz bir şey var mı? ',
            kvkk_col_name, # Add the exact KVKK column name
            tc_kimlik_col_name, # Add the exact TC Kimlik column name
            okul_col_name, # Add Okul original name
            bolum_col_name, # Add Bolum original name
            dogum_tarihi_col_name, # Add Dogum Tarihi original name
            '13. sütun', # Based on CSV header
            '12. sütun'  # Based on CSV header
        ]

        removed_cols_count = 0
        cols_before_removal = df.columns.tolist() # Get columns before removal for accurate checking

        print("Attempting to remove specified columns...")
        for col_name_to_check in list(columns_to_remove):
            found_col = None
            current_df_cols = [str(c) for c in df.columns]
            col_name_to_check_str = str(col_name_to_check)

            for df_col_str in current_df_cols:
                if df_col_str.strip() == col_name_to_check_str.strip():
                    for original_df_col in df.columns:
                         if str(original_df_col) == df_col_str:
                              found_col = original_df_col
                              break
                    break

            if found_col is not None and found_col in df.columns:
                try:
                    df = df.drop(columns=[found_col])
                    print(f" - Removed column: '{found_col}' (Matched based on '{col_name_to_check_str}')")
                    removed_cols_count += 1
                except KeyError:
                    print(f" - INFO: Column '{found_col}' was already removed or not found during drop.")
            else:
                was_in_list = False
                for item in columns_to_remove:
                     if str(item).strip() == col_name_to_check_str.strip():
                          was_in_list = True
                          break
                if was_in_list:
                     print(f" - WARNING: Column matching '{col_name_to_check_str}' not found. Skipping removal.")

        if removed_cols_count > 0:
            print(f"Successfully removed {removed_cols_count} columns.")

        # --- Reorder columns (UUID and Counter come first) ---
        cols = df.columns.tolist()
        if uuid_col in cols: cols.remove(uuid_col)
        if counter_col in cols: cols.remove(counter_col)
        df = df[[uuid_col, counter_col] + cols]
        print(f"Moved '{uuid_col}' and '{counter_col}' columns to the beginning.")
        
        # Rename column names (ensure these target names exist after removals)
        rename_map = {
            "Ad-Soyad": "isim",
            "E-posta adresiniz": "mail",
        }
        actual_rename_map = {}
        print("Attempting to rename columns...")
        for original_name, new_name in rename_map.items():
            found_original_col = None
            for df_col in df.columns:
                 if str(df_col).strip() == str(original_name).strip():
                     found_original_col = df_col
                     break
            if found_original_col:
                 actual_rename_map[found_original_col] = new_name
                 print(f" - Will rename '{found_original_col}' to '{new_name}'")
            else:
                 print(f" - WARNING: Column matching '{original_name}' not found for renaming.")

        df.rename(columns=actual_rename_map, inplace=True)
        if actual_rename_map:
            print("Finished renaming columns.")
        else:
            print("No columns were renamed.")

        # Convert email column to lowercase using Turkish-specific handling
        if 'mail' in df.columns:
            def turkish_lowercase(text):
                if pd.isna(text):
                    return text
                text = str(text)
                text = text.replace('\u0130', '\u0069')
                return text.lower()

            df['mail'] = df['mail'].apply(turkish_lowercase)
            print("Applied Turkish-specific lowercase conversion to 'mail' column.")
        else:
             print("Warning: 'mail' column not found after renaming. Cannot convert to lowercase.")

        # --- SAVE INTERMEDIATE CSV (FORM DATA) ---
        if not os.path.exists(CSV_OUTPUT_DIR):
            os.makedirs(CSV_OUTPUT_DIR)
            print(f"Created CSV output directory: '{CSV_OUTPUT_DIR}'")
        
        form_csv_filename = os.path.splitext(os.path.basename(file_path))[0] + '_form.csv'
        form_csv_path = os.path.join(CSV_OUTPUT_DIR, form_csv_filename)
        try:
            df.to_csv(form_csv_path, index=False, encoding='utf-8-sig')
            print(f"Successfully saved intermediate form data CSV to '{form_csv_path}'.")
        except Exception as e_form_csv:
            print(f"Warning: Could not save intermediate form data CSV: {e_form_csv}")
        # --- END SAVE INTERMEDIATE CSV ---

        # --- Handle Duplicates based on Timestamp ---
        timestamp_col_original_name = 'Zaman damgası'
        actual_timestamp_col = None
        for col in df.columns:
            if str(col).strip() == timestamp_col_original_name:
                actual_timestamp_col = col
                break

        if actual_timestamp_col:
            print(f"Found timestamp column: '{actual_timestamp_col}'. Processing duplicates...")
            try:
                df[actual_timestamp_col] = pd.to_datetime(df[actual_timestamp_col], errors='coerce')
                original_rows = len(df)
                df.dropna(subset=[actual_timestamp_col], inplace=True)
                if len(df) < original_rows:
                    print(f" - WARNING: Removed {original_rows - len(df)} rows due to invalid timestamp format.")

                df.sort_values(by=['mobile', actual_timestamp_col], ascending=[True, False], inplace=True)
                rows_before_dedup = len(df)
                df.drop_duplicates(subset=['mobile'], keep='first', inplace=True)
                rows_after_dedup = len(df)
                removed_duplicates = rows_before_dedup - rows_after_dedup
                if removed_duplicates > 0:
                    print(f" - Removed {removed_duplicates} older duplicate entries based on phone number.")
                else:
                    print(" - No duplicate phone numbers found to remove.")

            except Exception as e:
                print(f" - WARNING: Could not process duplicates due to error: {e}. Skipping deduplication.")
        else:
            print(f" - WARNING: Timestamp column '{timestamp_col_original_name}' not found. Cannot remove duplicates based on time.")

        # --- CSV output ---
        # Change filename to include _clean
        output_csv_filename = os.path.splitext(os.path.basename(file_path))[0] + '_clean.csv'
        output_csv_path = os.path.join(CSV_OUTPUT_DIR, output_csv_filename)
        df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
        print(f"Successfully saved final processed CSV to '{output_csv_path}'.")
        return output_csv_path
    except Exception as e:
        print(f"An unexpected error occurred during Excel processing: {e}")
        import traceback
        traceback.print_exc()
        return None

# --- QR Code Generation ---
from QRGenerator import generate_qr_codes_from_csv

# --- Excel with QR Code Generation ---
def generate_excel_with_qr(csv_path, qr_dir, excel_output_path):
    df = pd.read_csv(csv_path, dtype=str)
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Kodlar"
    ws.column_dimensions['A'].width = 20  # Column width for QR
    # Create header row
    headers = ['qr kodlar', 'ad soyad', 'posta', 'numara']
    ws.append(headers)
    row_number = 2
    for _, row in df.iterrows():
        ws.row_dimensions[row_number].height = 100  # Adjust row height according to QR size
        mobile = row.get('mobile', '').strip() # Ensure mobile is stripped

        img_path = None
        if mobile: # Only proceed if mobile is not empty
            img_path = os.path.join(qr_dir, f"{mobile}.png") # Use mobile directly for filename

        # Other columns
        ws.cell(row=row_number, column=2, value=row.get('isim', ''))
        ws.cell(row=row_number, column=3, value=row.get('mail', ''))
        ws.cell(row=row_number, column=4, value=row.get('mobile', ''))
        # Add QR image (if exists)
        if img_path and os.path.exists(img_path):
            img = OpenpyxlImage(img_path)
            img.width = 145
            img.height = 145
            ws.add_image(img, f"A{row_number}")
        else:
            ws.cell(row=row_number, column=1, value="No QR")
            if mobile:
                 print(f"Warning: QR image not found for mobile '{mobile}' at expected path: {img_path}")
        row_number += 1

    # Adjust other column widths
    for col in ['B', 'C', 'D']:
        max_length = 0
        for cell in ws[col]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col].width = max_length + 5  # add padding

    # Ensure the directory exists before saving
    excel_dir = os.path.dirname(excel_output_path)
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    wb.save(excel_output_path)
    print(f"Successfully generated Excel with QR codes at: '{excel_output_path}'")


# --- File Operations ---
# from FileOperations import create_directory_if_not_exists # This line is now handled above

# --- Main Execution ---
if __name__ == "__main__":
    # --- Find Input Excel File ---
    if not os.path.exists(INPUT_DIR):
        os.makedirs(INPUT_DIR)
        print(f"Created directory '{INPUT_DIR}'. Please place your single Excel (.xlsx) file inside it and rerun the script.")
        sys.exit(1)

    excel_files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx') and not f.startswith('~')] # Ignore temp files

    if len(excel_files) == 0:
        print(f"Error: No Excel (.xlsx) file found in the '{INPUT_DIR}' directory.")
        print("Please place the input Excel file there.")
        sys.exit(1)
    elif len(excel_files) > 1:
        print(f"Error: Multiple Excel (.xlsx) files found in the '{INPUT_DIR}' directory:")
        for f in excel_files:
            print(f" - {f}")
        print("Please ensure only one Excel file is present in the input directory.")
        sys.exit(1)
    else:
        excel_file_path = os.path.join(INPUT_DIR, excel_files[0])
        safe_excel_file_path_repr = repr(excel_file_path.encode(sys.stdout.encoding, errors='replace').decode(sys.stdout.encoding, errors='replace'))
        print(f"Using input file: {safe_excel_file_path_repr}")

    # Define output directory paths
    designed_qr_output_dir = os.path.join('output', 'designed_qr')
    excel_output_dir = os.path.join('output', 'excel')

    csv_file = process_excel(excel_file_path, PHONE_COLUMN_NAME, UUID_COLUMN_NAME, COUNTER_COLUMN_NAME)
    if csv_file: # csv_file now holds the path to *_clean.csv
        create_directory_if_not_exists(QR_OUTPUT_DIR)
        create_directory_if_not_exists(excel_output_dir)
        create_directory_if_not_exists(designed_qr_output_dir)

        # --- QR code generation and Excel ---
        # Pass the correct csv_file path (_clean.csv)
        generate_qr_codes_from_csv(csv_file, UUID_COLUMN_NAME, 'mobile', QR_OUTPUT_DIR)
        print("QR code generation completed.")

        base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
        dynamic_excel_output_path = os.path.join(excel_output_dir, f"{base_name}_modified.xlsx")

        # Pass the correct csv_file path (_clean.csv)
        generate_excel_with_qr(csv_file, QR_OUTPUT_DIR, dynamic_excel_output_path)
        print("Excel generation with QR completed.") # Corrected print message

        # --- QR Design ---
        print("\n--- Starting QR Design Process ---")
        template_image_path = "tasarim.jpg"
        if not os.path.exists(template_image_path):
             print(f"Warning: Template image '{template_image_path}' not found. Skipping QR design.")
             can_design = False
        else:
            # Pass the correct csv_file path (_clean.csv)
            overlay_qr_on_template(QR_OUTPUT_DIR, template_image_path, designed_qr_output_dir, csv_path=csv_file)
            print("QR Design process completed.")
            can_design = True

        # --- Firebase Sync ---
        firebase_choice = input("Do you want to sync with Firebase? (yes/no): ").strip().lower()
        if firebase_choice == 'yes':
            print("\n--- Starting Firebase Synchronization ---")
            try:
                if not os.path.exists(FIREBASE_SYNC_SCRIPT):
                    print(f"Error: Firebase sync script '{FIREBASE_SYNC_SCRIPT}' not found.")
                else:
                    python_executable = sys.executable
                    # Pass the correct csv_file path (_clean.csv) to the sync script
                    result = subprocess.run(
                        [python_executable, FIREBASE_SYNC_SCRIPT, csv_file],
                        capture_output=True, text=True, check=True
                    )
                    print("Firebase sync script output:")
                    print(result.stdout)
                    if result.stderr: print("Firebase sync script errors:", result.stderr)
                    print("--- Firebase Synchronization Finished ---")
            except FileNotFoundError:
                print(f"Error: Python executable not found at '{sys.executable}'. Cannot run sync script.")
            except subprocess.CalledProcessError as e:
                print(f"Error: Firebase sync script failed with exit code {e.returncode}.")
                print("Output:")
                print(e.stdout)
                print("Errors:")
                print(e.stderr)
            except Exception as e:
                print(f"An unexpected error occurred during Firebase sync: {e}")

        # --- Send QR Emails ---
        email_choice = input("Do you want to send emails with designed QR codes? (yes/no): ").strip().lower()
        if email_choice == 'yes':
            # Check existence of the correct csv_file (_clean.csv)
            if can_design and os.path.exists(csv_file) and os.path.exists(designed_qr_output_dir):
                print("\n--- Starting QR Email Sending Process ---")
                # Pass the correct csv_file path (_clean.csv)
                send_qr_codes(
                    csv_file, designed_qr_output_dir, SENDER_EMAIL,
                    SENDER_PASSWORD, SMTP_SERVER, SMTP_PORT
                )
                print("--- QR Email Sending Process Finished ---")
            else:
                # Update the check message to reflect the correct csv file name pattern
                print(f"Prerequisites not met (Template exists: {can_design}, CSV exists: {os.path.exists(csv_file)}, Designed QRs exist: {os.path.exists(designed_qr_output_dir)}). QR Emails cannot be sent.")

        print("\nAll processes have been completed.")
        print("Note: To generate and send certificates, run 'python CertificateGeneratorSender.py' separately.")
    else:
        print("CSV file generation failed. Skipping subsequent steps.")
